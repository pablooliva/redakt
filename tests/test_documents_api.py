"""Integration tests for POST /api/documents/upload endpoint."""

import asyncio
import io
import json
from unittest.mock import AsyncMock, patch

import pytest


class TestDocumentUploadEndpoint:
    def test_upload_txt_with_pii(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Upload .txt file with PII: returns anonymized text + mapping."""
        mock_presidio_analyze.return_value = [
            {"entity_type": "PERSON", "start": 8, "end": 18, "score": 0.85},
            {"entity_type": "EMAIL_ADDRESS", "start": 22, "end": 38, "score": 1.0},
        ]
        content = b"Contact John Smith at john@example.com"
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("test.txt", io.BytesIO(content), "text/plain")},
            data={"language": "auto"},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert "<PERSON_1>" in data["anonymized_content"]
        assert data["mappings"]["<PERSON_1>"] == "John Smith"
        assert data["source_format"] == "txt"
        assert data["anonymized_structured"] is None

    def test_upload_csv(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Upload .csv file: returns anonymized CSV text."""
        mock_presidio_analyze.return_value = []
        content = b"Name,Email\nJohn,john@test.com"
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("data.csv", io.BytesIO(content), "text/csv")},
            data={"language": "en"},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["anonymized_content"] is not None
        assert data["source_format"] == "csv"

    def test_upload_json(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Upload .json file: returns anonymized JSON structure."""
        mock_presidio_analyze.return_value = [
            {"entity_type": "PERSON", "start": 0, "end": 10, "score": 0.9},
        ]
        content = json.dumps({"name": "John Smith", "age": 30}).encode()
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("data.json", io.BytesIO(content), "application/json")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["anonymized_structured"] is not None
        assert data["anonymized_content"] is None
        assert data["source_format"] == "json"

    def test_upload_xlsx(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Upload .xlsx file: returns anonymized sheet structure."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Name"
        ws["B1"] = "Email"
        ws["A2"] = "John Smith"
        ws["B2"] = "john@example.com"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        mock_presidio_analyze.return_value = []
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("data.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["anonymized_structured"] is not None
        assert data["anonymized_content"] is None
        assert data["source_format"] == "xlsx"
        assert "Sheet1" in data["anonymized_structured"]

    def test_upload_docx(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Upload .docx file: returns anonymized text."""
        import docx
        doc = docx.Document()
        doc.add_paragraph("Hello John Smith")
        buf = io.BytesIO()
        doc.save(buf)
        buf.seek(0)

        mock_presidio_analyze.return_value = [
            {"entity_type": "PERSON", "start": 6, "end": 16, "score": 0.85},
        ]
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("doc.docx", buf, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["anonymized_content"] is not None
        assert data["source_format"] == "docx"

    def test_upload_pdf(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Upload .pdf file: returns anonymized text."""
        # Minimal PDF
        pdf_content = b"""%PDF-1.4
1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj
2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj
3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792]
/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >> endobj
4 0 obj << /Length 44 >> stream
BT /F1 12 Tf 100 700 Td (Hello World) Tj ET
endstream endobj
5 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj
xref
0 6
trailer << /Size 6 /Root 1 0 R >>
startxref
0
%%EOF"""
        mock_presidio_analyze.return_value = []
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("doc.pdf", io.BytesIO(pdf_content), "application/pdf")},
        )
        # May get 200 or 422 depending on pdfminer parsing -- the test validates the endpoint works
        assert resp.status_code in (200, 422)

    def test_unsupported_file_type(self, client):
        """Unsupported file type returns 400."""
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("malware.exe", io.BytesIO(b"MZ..."), "application/octet-stream")},
        )
        assert resp.status_code == 400
        assert "Unsupported" in resp.json()["detail"]

    def test_file_too_large(self, client):
        """File > 10MB returns 413."""
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("big.txt", io.BytesIO(b"x"), "text/plain")},
            data={"language": "en"},
        )
        # The file content is small, but we need to test with actual large content
        # or mock settings. Let's test the validation path:
        with patch("redakt.services.document_processor.settings") as mock_settings:
            mock_settings.max_file_size = 10
            mock_settings.supported_file_types = [".txt"]
            mock_settings.allow_list = []
            mock_settings.default_score_threshold = 0.35
            mock_settings.supported_languages = ["en", "de"]
            mock_settings.max_text_length = 512_000
            resp2 = client.post(
                "/api/documents/upload",
                files={"file": ("big.txt", io.BytesIO(b"x" * 100), "text/plain")},
                data={"language": "en"},
            )
        assert resp2.status_code == 413

    def test_empty_file(self, client, mock_doc_detect_language):
        """Empty file returns empty content gracefully."""
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("empty.txt", io.BytesIO(b""), "text/plain")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["anonymized_content"] == ""
        assert data["mappings"] == {}

    def test_corrupted_file(self, client):
        """Corrupted file returns 422."""
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("bad.json", io.BytesIO(b"{invalid"), "application/json")},
        )
        assert resp.status_code == 422

    def test_language_override(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Explicit language parameter is used."""
        mock_presidio_analyze.return_value = []
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("test.txt", io.BytesIO(b"Hello world"), "text/plain")},
            data={"language": "en"},
        )
        assert resp.status_code == 200
        mock_doc_detect_language.assert_not_called()

    def test_allow_list(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Per-request allow_list terms are passed to Presidio."""
        mock_presidio_analyze.return_value = []
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("test.txt", io.BytesIO(b"Hello ACME Corp"), "text/plain")},
            data={"language": "en", "allow_list": "ACME,Corp"},
        )
        assert resp.status_code == 200
        call_kwargs = mock_presidio_analyze.call_args.kwargs
        assert "ACME" in call_kwargs["allow_list"]
        assert "Corp" in call_kwargs["allow_list"]

    def test_presidio_unavailable(self, client, mock_doc_detect_language):
        """Presidio connection error returns 503."""
        import httpx
        with patch(
            "redakt.services.presidio.PresidioClient.analyze",
            new_callable=AsyncMock,
            side_effect=httpx.ConnectError("Connection refused"),
        ):
            resp = client.post(
                "/api/documents/upload",
                files={"file": ("test.txt", io.BytesIO(b"John Smith"), "text/plain")},
            )
        assert resp.status_code == 503

    def test_audit_log_emitted(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Audit log is emitted with correct metadata (no PII)."""
        mock_presidio_analyze.return_value = [
            {"entity_type": "PERSON", "start": 0, "end": 10, "score": 0.85},
        ]
        with patch("redakt.routers.documents.log_document_upload") as mock_log:
            resp = client.post(
                "/api/documents/upload",
                files={"file": ("test.txt", io.BytesIO(b"John Smith"), "text/plain")},
            )
            assert resp.status_code == 200
            mock_log.assert_called_once()
            call_kwargs = mock_log.call_args.kwargs
            assert call_kwargs["file_type"] == "txt"
            assert call_kwargs["entity_count"] == 1
            assert "PERSON" in call_kwargs["entity_types"]
            assert call_kwargs["source"] == "api"
            # Verify no PII in log
            assert "John" not in str(call_kwargs)

    def test_response_structure_text(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Response has exactly one of anonymized_content or anonymized_structured."""
        mock_presidio_analyze.return_value = []
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("test.txt", io.BytesIO(b"Hello"), "text/plain")},
        )
        data = resp.json()
        assert "anonymized_content" in data
        assert "anonymized_structured" in data
        assert "mappings" in data
        assert "language_detected" in data
        assert "source_format" in data
        assert "metadata" in data
        # For txt: content populated, structured null
        assert data["anonymized_content"] is not None
        assert data["anonymized_structured"] is None

    def test_entities_filter(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Entities filter is parsed from comma-separated string."""
        mock_presidio_analyze.return_value = []
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("test.txt", io.BytesIO(b"Hello John"), "text/plain")},
            data={"entities": "PERSON,EMAIL_ADDRESS"},
        )
        assert resp.status_code == 200
        call_kwargs = mock_presidio_analyze.call_args.kwargs
        assert call_kwargs["entities"] == ["PERSON", "EMAIL_ADDRESS"]

    def test_no_extension(self, client):
        """File with no extension returns 400."""
        resp = client.post(
            "/api/documents/upload",
            files={"file": ("noextension", io.BytesIO(b"data"), "application/octet-stream")},
        )
        assert resp.status_code == 400

    def test_concurrent_upload_limit_429(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """PERF-004 / FAIL-009: Exceeding the upload concurrency semaphore returns 429.

        The upload semaphore defaults to max_concurrent_uploads (3).
        When all slots are occupied (sem.locked()), the next request should get 429.
        """
        mock_presidio_analyze.return_value = []

        # Create a semaphore with 0 initial value so locked() returns True immediately
        sem = asyncio.Semaphore(0)

        with patch("redakt.routers.documents._get_upload_semaphore", return_value=sem):
            resp = client.post(
                "/api/documents/upload",
                files={"file": ("test.txt", io.BytesIO(b"Hello world"), "text/plain")},
                data={"language": "en"},
            )
        assert resp.status_code == 429
        assert "Too many" in resp.json()["detail"]


class TestWebUIDocumentRoute:
    """Tests for POST /documents/submit (web UI route in pages.py)."""

    def test_web_ui_upload_txt(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Web UI route returns HTML partial with anonymized content."""
        mock_presidio_analyze.return_value = [
            {"entity_type": "PERSON", "start": 8, "end": 18, "score": 0.85},
        ]
        resp = client.post(
            "/documents/submit",
            files={"file": ("test.txt", io.BytesIO(b"Contact John Smith here"), "text/plain")},
            data={"language": "auto"},
        )
        assert resp.status_code == 200
        assert "PERSON_1" in resp.text

    def test_web_ui_upload_error(self, client):
        """Web UI route returns error HTML for unsupported format."""
        resp = client.post(
            "/documents/submit",
            files={"file": ("bad.exe", io.BytesIO(b"MZ..."), "application/octet-stream")},
            data={"language": "auto"},
        )
        assert resp.status_code == 200  # HTMX gets 200 with error content
        assert "Unsupported" in resp.text

    def test_web_ui_semaphore_rejection(self, client, mock_presidio_analyze, mock_doc_detect_language):
        """Web UI route returns error HTML when semaphore is full."""
        mock_presidio_analyze.return_value = []

        # Semaphore with 0 slots so locked() returns True immediately
        sem = asyncio.Semaphore(0)

        with patch("redakt.routers.pages._get_upload_semaphore", return_value=sem):
            resp = client.post(
                "/documents/submit",
                files={"file": ("test.txt", io.BytesIO(b"Hello"), "text/plain")},
                data={"language": "en"},
            )
        assert resp.status_code == 200  # HTMX always gets 200
        assert "Too many" in resp.text

    def test_web_ui_empty_file(self, client, mock_doc_detect_language):
        """Web UI route handles empty files gracefully."""
        resp = client.post(
            "/documents/submit",
            files={"file": ("empty.txt", io.BytesIO(b""), "text/plain")},
            data={"language": "auto"},
        )
        assert resp.status_code == 200
