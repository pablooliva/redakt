"""Unit tests for src/redakt/services/document_processor.py"""

from unittest.mock import AsyncMock, patch

import pytest

from redakt.services.document_processor import (
    DocumentProcessingError,
    build_unified_placeholder_map,
    process_document,
    validate_file,
)
from redakt.services.extractors import TextChunk


# ---------------------------------------------------------------------------
# build_unified_placeholder_map tests
# ---------------------------------------------------------------------------

class TestBuildUnifiedPlaceholderMap:
    def test_single_chunk(self):
        all_chunk_entities = [
            [
                {"entity_type": "PERSON", "start": 0, "end": 10, "score": 0.9, "original_text": "John Smith"},
                {"entity_type": "EMAIL_ADDRESS", "start": 20, "end": 36, "score": 1.0, "original_text": "john@example.com"},
            ]
        ]
        mappings, per_chunk = build_unified_placeholder_map(all_chunk_entities)
        assert mappings == {
            "<PERSON_1>": "John Smith",
            "<EMAIL_ADDRESS_1>": "john@example.com",
        }
        assert per_chunk[0][0] == "<PERSON_1>"
        assert per_chunk[0][1] == "<EMAIL_ADDRESS_1>"

    def test_same_pii_across_chunks(self):
        """Same PII value in different chunks -> same placeholder."""
        all_chunk_entities = [
            [{"entity_type": "PERSON", "start": 0, "end": 10, "score": 0.9, "original_text": "John Smith"}],
            [{"entity_type": "PERSON", "start": 0, "end": 10, "score": 0.9, "original_text": "John Smith"}],
        ]
        mappings, per_chunk = build_unified_placeholder_map(all_chunk_entities)
        assert mappings == {"<PERSON_1>": "John Smith"}
        assert per_chunk[0][0] == "<PERSON_1>"
        assert per_chunk[1][0] == "<PERSON_1>"

    def test_different_pii_values(self):
        all_chunk_entities = [
            [{"entity_type": "PERSON", "start": 0, "end": 10, "score": 0.9, "original_text": "John Smith"}],
            [{"entity_type": "PERSON", "start": 0, "end": 8, "score": 0.9, "original_text": "Jane Doe"}],
        ]
        mappings, per_chunk = build_unified_placeholder_map(all_chunk_entities)
        assert mappings == {
            "<PERSON_1>": "John Smith",
            "<PERSON_2>": "Jane Doe",
        }

    def test_counter_per_type_across_chunks(self):
        all_chunk_entities = [
            [{"entity_type": "PERSON", "start": 0, "end": 10, "score": 0.9, "original_text": "John Smith"}],
            [
                {"entity_type": "EMAIL_ADDRESS", "start": 0, "end": 16, "score": 1.0, "original_text": "john@example.com"},
                {"entity_type": "PERSON", "start": 20, "end": 28, "score": 0.9, "original_text": "Jane Doe"},
            ],
        ]
        mappings, per_chunk = build_unified_placeholder_map(all_chunk_entities)
        assert "<PERSON_1>" in mappings
        assert "<PERSON_2>" in mappings
        assert "<EMAIL_ADDRESS_1>" in mappings
        assert mappings["<PERSON_1>"] == "John Smith"
        assert mappings["<PERSON_2>"] == "Jane Doe"

    def test_empty_chunks(self):
        mappings, per_chunk = build_unified_placeholder_map([[], []])
        assert mappings == {}
        assert per_chunk == [{}, {}]


# ---------------------------------------------------------------------------
# validate_file tests
# ---------------------------------------------------------------------------

class TestValidateFile:
    def test_rejects_oversized(self):
        with pytest.raises(DocumentProcessingError, match="exceeds the maximum size"):
            validate_file(b"x" * 100, ".txt", 11 * 1024 * 1024)

    def test_rejects_unsupported_extension(self):
        with pytest.raises(DocumentProcessingError, match="Unsupported file format"):
            validate_file(b"data", ".exe", 100)

    def test_rejects_no_extension(self):
        with pytest.raises(DocumentProcessingError, match="Unsupported file format"):
            validate_file(b"data", "", 100)

    def test_rejects_magic_byte_mismatch(self):
        with pytest.raises(DocumentProcessingError, match="does not match"):
            validate_file(b"not a pdf", ".pdf", 9)

    def test_accepts_valid_pdf(self):
        validate_file(b"%PDF-1.4 content", ".pdf", 16)

    def test_accepts_valid_txt(self):
        validate_file(b"hello world", ".txt", 11)

    def test_accepts_valid_xlsx(self):
        validate_file(b"PK\x03\x04 rest", ".xlsx", 8)


# ---------------------------------------------------------------------------
# process_document tests (integration with mocked Presidio)
# ---------------------------------------------------------------------------

class TestProcessDocument:
    @pytest.fixture
    def mock_presidio(self):
        presidio = AsyncMock()
        presidio.analyze = AsyncMock(return_value=[])
        return presidio

    @pytest.fixture
    def mock_lang(self):
        with patch(
            "redakt.services.document_processor.detect_language",
            new_callable=AsyncMock,
        ) as mock:
            mock.return_value = "en"
            yield mock

    @pytest.mark.asyncio
    async def test_txt_file_no_pii(self, mock_presidio, mock_lang):
        raw = b"The weather is nice today."
        result = await process_document(
            raw=raw, extension=".txt", file_size=len(raw), presidio=mock_presidio
        )
        assert result["anonymized_content"] == "The weather is nice today."
        assert result["mappings"] == {}
        assert result["source_format"] == "txt"

    @pytest.mark.asyncio
    async def test_txt_file_with_pii(self, mock_presidio, mock_lang):
        raw = b"Contact John Smith at john@example.com"
        mock_presidio.analyze.return_value = [
            {"entity_type": "PERSON", "start": 8, "end": 18, "score": 0.85},
            {"entity_type": "EMAIL_ADDRESS", "start": 22, "end": 38, "score": 1.0},
        ]
        result = await process_document(
            raw=raw, extension=".txt", file_size=len(raw), presidio=mock_presidio
        )
        assert "<PERSON_1>" in result["anonymized_content"]
        assert "<EMAIL_ADDRESS_1>" in result["anonymized_content"]
        assert result["mappings"]["<PERSON_1>"] == "John Smith"

    @pytest.mark.asyncio
    async def test_empty_file(self, mock_presidio, mock_lang):
        result = await process_document(
            raw=b"", extension=".txt", file_size=0, presidio=mock_presidio
        )
        assert result["anonymized_content"] == ""
        assert result["mappings"] == {}

    @pytest.mark.asyncio
    async def test_oversized_chunk_skipped(self, mock_presidio, mock_lang):
        """Chunk > max_text_length should be replaced with placeholder."""
        large_text = "x" * 600_000  # > 512KB
        raw = large_text.encode("utf-8")
        with patch("redakt.services.document_processor.settings") as mock_settings:
            mock_settings.max_text_length = 512_000
            mock_settings.max_file_size = 10 * 1024 * 1024
            mock_settings.supported_file_types = [".txt"]
            mock_settings.allow_list = []
            mock_settings.default_score_threshold = 0.35
            mock_settings.supported_languages = ["en", "de"]
            result = await process_document(
                raw=raw, extension=".txt", file_size=len(raw), presidio=mock_presidio
            )
        assert result["anonymized_content"] == "[CONTENT TOO LARGE - SKIPPED]"
        assert any("too large" in w.lower() or "skipped" in w.lower()
                    for w in result["metadata"]["warnings"])

    @pytest.mark.asyncio
    async def test_language_detection_called(self, mock_presidio, mock_lang):
        raw = b"Some text for analysis"
        await process_document(
            raw=raw, extension=".txt", file_size=len(raw), presidio=mock_presidio
        )
        mock_lang.assert_called_once()

    @pytest.mark.asyncio
    async def test_language_explicit(self, mock_presidio, mock_lang):
        raw = b"Some text"
        await process_document(
            raw=raw, extension=".txt", file_size=len(raw),
            presidio=mock_presidio, language="en"
        )
        mock_lang.assert_not_called()

    @pytest.mark.asyncio
    async def test_json_structured_output(self, mock_presidio, mock_lang):
        data = {"name": "John Smith", "count": 42}
        raw = json.dumps(data).encode()
        mock_presidio.analyze.return_value = [
            {"entity_type": "PERSON", "start": 0, "end": 10, "score": 0.9},
        ]
        result = await process_document(
            raw=raw, extension=".json", file_size=len(raw), presidio=mock_presidio
        )
        assert result["anonymized_content"] is None
        assert result["anonymized_structured"] is not None
        assert result["anonymized_structured"]["count"] == 42
        assert "<PERSON_1>" in str(result["anonymized_structured"]["name"])

    @pytest.mark.asyncio
    async def test_csv_output(self, mock_presidio, mock_lang):
        raw = b"Name,Email\nJohn,john@test.com"
        mock_presidio.analyze.return_value = []
        result = await process_document(
            raw=raw, extension=".csv", file_size=len(raw), presidio=mock_presidio
        )
        assert result["anonymized_content"] is not None
        assert result["anonymized_structured"] is None
        assert "Name" in result["anonymized_content"]

    @pytest.mark.asyncio
    async def test_rejects_unsupported_format(self, mock_presidio, mock_lang):
        with pytest.raises(DocumentProcessingError, match="Unsupported"):
            await process_document(
                raw=b"data", extension=".exe", file_size=4, presidio=mock_presidio
            )

    @pytest.mark.asyncio
    async def test_rejects_oversized_file(self, mock_presidio, mock_lang):
        with pytest.raises(DocumentProcessingError, match="exceeds"):
            await process_document(
                raw=b"x", extension=".txt", file_size=11 * 1024 * 1024,
                presidio=mock_presidio
            )


import io
import json


# ---------------------------------------------------------------------------
# Multi-sheet PII consistency (EDGE-006)
# ---------------------------------------------------------------------------

class TestMultiSheetPiiConsistency:
    """EDGE-006: Same PII in different sheets produces same placeholder."""

    @pytest.fixture
    def mock_presidio(self):
        presidio = AsyncMock()
        return presidio

    @pytest.fixture
    def mock_lang(self):
        with patch(
            "redakt.services.document_processor.detect_language",
            new_callable=AsyncMock,
        ) as mock:
            mock.return_value = "en"
            yield mock

    @pytest.mark.asyncio
    async def test_same_pii_across_sheets(self, mock_presidio, mock_lang):
        """Same name in Sheet1 and Sheet2 gets the same placeholder."""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws1 = wb.create_sheet("Sheet1")
        ws1["A1"] = "John Smith"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "John Smith"
        buf = io.BytesIO()
        wb.save(buf)
        raw = buf.getvalue()

        # Mock Presidio to detect PERSON in both sheets
        def analyze_side_effect(**kwargs):
            text = kwargs.get("text", "")
            if text == "John Smith":
                return [{"entity_type": "PERSON", "start": 0, "end": 10, "score": 0.9}]
            return []

        mock_presidio.analyze = AsyncMock(side_effect=analyze_side_effect)

        result = await process_document(
            raw=raw, extension=".xlsx", file_size=len(raw), presidio=mock_presidio
        )
        # Both sheets should use <PERSON_1>
        assert result["mappings"] == {"<PERSON_1>": "John Smith"}
        sheet1 = result["anonymized_structured"]["Sheet1"]
        sheet2 = result["anonymized_structured"]["Sheet2"]
        assert sheet1[0][0] == "<PERSON_1>"
        assert sheet2[0][0] == "<PERSON_1>"
